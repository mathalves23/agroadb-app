"""
Sistema de Geração de Relatórios (PDF e Excel)
Gera relatórios profissionais de investigações
"""
from typing import Optional, Dict, Any, List
from datetime import datetime
from io import BytesIO
import logging

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ReportGenerator:
    """
    Gerador de Relatórios Profissionais
    
    Suporta PDF e Excel com templates customizáveis
    """
    
    def __init__(self):
        self.logo_path = None  # TODO: Adicionar logo da empresa
    
    # ==================== PDF Generation ====================
    
    def generate_pdf_report(
        self,
        investigation: Dict[str, Any],
        results: Dict[str, Any],
        report_type: str = "detailed"  # "detailed" ou "executive"
    ) -> BytesIO:
        """
        Gera relatório em PDF
        
        Args:
            investigation: Dados da investigação
            results: Resultados dos scrapers
            report_type: Tipo de relatório ("detailed" ou "executive")
            
        Returns:
            BytesIO com o PDF
        """
        buffer = BytesIO()
        
        # Criar documento
        if report_type == "executive":
            doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=18)
        else:
            doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                                   topMargin=72, bottomMargin=18)
        
        # Container para elementos
        elements = []
        
        # Estilos
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='Center', alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='Right', alignment=TA_RIGHT))
        
        # Cabeçalho
        elements.extend(self._create_pdf_header(investigation, styles))
        
        # Sumário Executivo
        elements.extend(self._create_pdf_executive_summary(investigation, results, styles))
        
        if report_type == "detailed":
            # Seções detalhadas
            elements.extend(self._create_pdf_detailed_sections(results, styles))
        
        # Rodapé
        elements.extend(self._create_pdf_footer(styles))
        
        # Construir PDF
        doc.build(elements)
        
        buffer.seek(0)
        return buffer
    
    def _create_pdf_header(self, investigation: Dict[str, Any], styles) -> List:
        """Cria cabeçalho do PDF"""
        elements = []
        
        # Logo (se disponível)
        if self.logo_path:
            try:
                logo = Image(self.logo_path, width=2*inch, height=0.5*inch)
                elements.append(logo)
            except:
                pass
        
        # Título
        title = Paragraph(
            f"<b>RELATÓRIO DE INVESTIGAÇÃO PATRIMONIAL</b>",
            styles['Title']
        )
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Informações da investigação
        info_data = [
            ['ID da Investigação:', investigation.get('id', 'N/A')],
            ['Alvo:', investigation.get('target_name', 'N/A')],
            ['CPF/CNPJ:', investigation.get('target_cpf_cnpj', 'N/A')],
            ['Data:', datetime.utcnow().strftime('%d/%m/%Y')],
            ['Status:', investigation.get('status', 'N/A')]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*inch))
        
        return elements
    
    def _create_pdf_executive_summary(
        self,
        investigation: Dict[str, Any],
        results: Dict[str, Any],
        styles
    ) -> List:
        """Cria sumário executivo"""
        elements = []
        
        # Título da seção
        elements.append(Paragraph("<b>1. SUMÁRIO EXECUTIVO</b>", styles['Heading1']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Estatísticas gerais
        total_properties = len(results.get('car', [])) + len(results.get('incra', []))
        total_companies = len(results.get('receita', []))
        total_documents = len(results.get('diario_oficial', []))
        
        summary_text = f"""
        <para>
        Esta investigação patrimonial identificou <b>{total_properties} propriedades</b>, 
        <b>{total_companies} empresas</b> e <b>{total_documents} menções</b> em diários oficiais 
        relacionadas ao alvo investigado.
        </para>
        """
        
        elements.append(Paragraph(summary_text, styles['BodyText']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabela de resultados por fonte
        summary_data = [
            ['Fonte', 'Resultados'],
            ['CAR (Propriedades Rurais)', str(len(results.get('car', [])))],
            ['INCRA (SNCR)', str(len(results.get('incra', [])))],
            ['Receita Federal', str(len(results.get('receita', [])))],
            ['Diários Oficiais', str(len(results.get('diario_oficial', [])))],
            ['Cartórios', str(len(results.get('cartorios', [])))],
            ['SIGEF/SICAR', str(len(results.get('sigef_sicar', [])))]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(PageBreak())
        
        return elements
    
    def _create_pdf_detailed_sections(self, results: Dict[str, Any], styles) -> List:
        """Cria seções detalhadas do relatório"""
        elements = []
        
        # Seção 2: Propriedades Rurais
        if results.get('car') or results.get('incra'):
            elements.append(Paragraph("<b>2. PROPRIEDADES RURAIS</b>", styles['Heading1']))
            elements.append(Spacer(1, 0.2*inch))
            
            # CAR
            if results.get('car'):
                elements.append(Paragraph("<b>2.1 Cadastro Ambiental Rural (CAR)</b>", styles['Heading2']))
                elements.append(Spacer(1, 0.1*inch))
                
                for idx, prop in enumerate(results['car'][:5], 1):  # Primeiras 5
                    prop_text = f"""
                    <b>Propriedade {idx}:</b> {prop.get('codigo_imovel', 'N/A')}<br/>
                    <b>Município:</b> {prop.get('municipio', 'N/A')} - {prop.get('estado', 'N/A')}<br/>
                    <b>Área:</b> {prop.get('area_hectares', 'N/A')} hectares<br/>
                    <b>Status:</b> {prop.get('status', 'N/A')}
                    """
                    elements.append(Paragraph(prop_text, styles['BodyText']))
                    elements.append(Spacer(1, 0.1*inch))
        
        # Seção 3: Vínculos Societários
        if results.get('receita'):
            elements.append(PageBreak())
            elements.append(Paragraph("<b>3. VÍNCULOS SOCIETÁRIOS</b>", styles['Heading1']))
            elements.append(Spacer(1, 0.2*inch))
            
            for idx, company in enumerate(results['receita'][:5], 1):
                company_text = f"""
                <b>Empresa {idx}:</b> {company.get('razao_social', 'N/A')}<br/>
                <b>CNPJ:</b> {company.get('cnpj', 'N/A')}<br/>
                <b>Situação:</b> {company.get('situacao', 'N/A')}<br/>
                <b>Capital Social:</b> R$ {company.get('capital_social', 'N/A')}
                """
                elements.append(Paragraph(company_text, styles['BodyText']))
                elements.append(Spacer(1, 0.1*inch))
        
        return elements
    
    def _create_pdf_footer(self, styles) -> List:
        """Cria rodapé do PDF"""
        elements = []
        
        elements.append(PageBreak())
        elements.append(Spacer(1, 0.5*inch))
        
        footer_text = """
        <para align=center>
        <b>OBSERVAÇÕES IMPORTANTES</b><br/><br/>
        Este relatório foi gerado automaticamente pela plataforma AgroADB.<br/>
        Os dados são obtidos de fontes públicas governamentais.<br/>
        A veracidade das informações deve ser confirmada nas fontes originais.<br/><br/>
        © 2026 AgroADB - Todos os direitos reservados
        </para>
        """
        
        elements.append(Paragraph(footer_text, styles['BodyText']))
        
        return elements
    
    # ==================== Excel Generation ====================
    
    def generate_excel_report(
        self,
        investigation: Dict[str, Any],
        results: Dict[str, Any]
    ) -> BytesIO:
        """
        Gera relatório em Excel
        
        Args:
            investigation: Dados da investigação
            results: Resultados dos scrapers
            
        Returns:
            BytesIO com o Excel
        """
        buffer = BytesIO()
        
        # Criar workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remover sheet padrão
        
        # Sheet 1: Sumário
        self._create_excel_summary_sheet(wb, investigation, results)
        
        # Sheet 2: CAR
        if results.get('car'):
            self._create_excel_car_sheet(wb, results['car'])
        
        # Sheet 3: INCRA
        if results.get('incra'):
            self._create_excel_incra_sheet(wb, results['incra'])
        
        # Sheet 4: Receita Federal
        if results.get('receita'):
            self._create_excel_receita_sheet(wb, results['receita'])
        
        # Sheet 5: Diários Oficiais
        if results.get('diario_oficial'):
            self._create_excel_diario_sheet(wb, results['diario_oficial'])
        
        # Salvar
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _create_excel_summary_sheet(
        self,
        wb: openpyxl.Workbook,
        investigation: Dict[str, Any],
        results: Dict[str, Any]
    ):
        """Cria sheet de sumário"""
        ws = wb.create_sheet("Sumário")
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Cabeçalho
        ws['A1'] = "RELATÓRIO DE INVESTIGAÇÃO PATRIMONIAL"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Informações
        ws['A3'] = "ID da Investigação:"
        ws['B3'] = investigation.get('id', 'N/A')
        
        ws['A4'] = "Alvo:"
        ws['B4'] = investigation.get('target_name', 'N/A')
        
        ws['A5'] = "CPF/CNPJ:"
        ws['B5'] = investigation.get('target_cpf_cnpj', 'N/A')
        
        ws['A6'] = "Data:"
        ws['B6'] = datetime.utcnow().strftime('%d/%m/%Y %H:%M')
        
        # Estatísticas
        ws['A8'] = "ESTATÍSTICAS"
        ws['A8'].font = header_font
        ws['A8'].fill = header_fill
        ws.merge_cells('A8:B8')
        
        stats = [
            ("CAR (Propriedades)", len(results.get('car', []))),
            ("INCRA (SNCR)", len(results.get('incra', []))),
            ("Receita Federal", len(results.get('receita', []))),
            ("Diários Oficiais", len(results.get('diario_oficial', []))),
            ("Cartórios", len(results.get('cartorios', []))),
            ("SIGEF/SICAR", len(results.get('sigef_sicar', [])))
        ]
        
        row = 9
        for label, count in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = count
            row += 1
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_excel_car_sheet(self, wb: openpyxl.Workbook, car_data: List[Dict]):
        """Cria sheet com dados do CAR"""
        ws = wb.create_sheet("CAR - Propriedades")
        
        # Cabeçalhos
        headers = ['Código', 'Município', 'Estado', 'Área (ha)', 'Status', 'Data Registro']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dados
        for row, prop in enumerate(car_data, 2):
            ws.cell(row=row, column=1).value = prop.get('codigo_imovel', '')
            ws.cell(row=row, column=2).value = prop.get('municipio', '')
            ws.cell(row=row, column=3).value = prop.get('estado', '')
            ws.cell(row=row, column=4).value = prop.get('area_hectares', 0)
            ws.cell(row=row, column=5).value = prop.get('status', '')
            ws.cell(row=row, column=6).value = prop.get('data_registro', '')
        
        # Ajustar largura
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_excel_incra_sheet(self, wb: openpyxl.Workbook, incra_data: List[Dict]):
        """Cria sheet com dados do INCRA"""
        ws = wb.create_sheet("INCRA - SNCR")
        
        # Cabeçalhos
        headers = ['CCIR', 'Município', 'Estado', 'Área (ha)', 'Módulos Fiscais', 'Situação']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dados
        for row, prop in enumerate(incra_data, 2):
            ws.cell(row=row, column=1).value = prop.get('ccir', '')
            ws.cell(row=row, column=2).value = prop.get('municipio', '')
            ws.cell(row=row, column=3).value = prop.get('estado', '')
            ws.cell(row=row, column=4).value = prop.get('area_total', 0)
            ws.cell(row=row, column=5).value = prop.get('modulos_fiscais', 0)
            ws.cell(row=row, column=6).value = prop.get('situacao', '')
        
        # Ajustar largura
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_excel_receita_sheet(self, wb: openpyxl.Workbook, receita_data: List[Dict]):
        """Cria sheet com dados da Receita Federal"""
        ws = wb.create_sheet("Receita Federal")
        
        # Cabeçalhos
        headers = ['CNPJ', 'Razão Social', 'Nome Fantasia', 'Situação', 'Capital Social', 'Abertura']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dados
        for row, company in enumerate(receita_data, 2):
            ws.cell(row=row, column=1).value = company.get('cnpj', '')
            ws.cell(row=row, column=2).value = company.get('razao_social', '')
            ws.cell(row=row, column=3).value = company.get('nome_fantasia', '')
            ws.cell(row=row, column=4).value = company.get('situacao', '')
            ws.cell(row=row, column=5).value = company.get('capital_social', 0)
            ws.cell(row=row, column=6).value = company.get('data_abertura', '')
        
        # Ajustar largura
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_excel_diario_sheet(self, wb: openpyxl.Workbook, diario_data: List[Dict]):
        """Cria sheet com diários oficiais"""
        ws = wb.create_sheet("Diários Oficiais")
        
        # Cabeçalhos
        headers = ['Data', 'Diário', 'Título', 'URL']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dados
        for row, pub in enumerate(diario_data, 2):
            ws.cell(row=row, column=1).value = pub.get('publication_date', '')
            ws.cell(row=row, column=2).value = pub.get('diary_name', '')
            ws.cell(row=row, column=3).value = pub.get('title', '')
            ws.cell(row=row, column=4).value = pub.get('url', '')
        
        # Ajustar largura
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40


# Instância global
report_generator = ReportGenerator()
