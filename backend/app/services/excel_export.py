"""
Excel Export Service for Investigations
"""
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO
import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from app.domain.investigation import Investigation
from app.domain.property import Property
from app.domain.company import Company
from app.domain.legal_query import LegalQuery


class ExcelExportService:
    """Service for exporting investigation data to Excel and CSV"""

    @staticmethod
    def generate_investigation_excel(
        investigation: Investigation,
        properties: List[Property],
        companies: List[Company],
        legal_queries: List[LegalQuery],
    ) -> BytesIO:
        """
        Generate Excel file with multiple sheets for investigation data
        
        Args:
            investigation: Investigation object
            properties: List of properties found
            companies: List of companies found
            legal_queries: List of legal queries executed
            
        Returns:
            BytesIO: Excel file in memory
        """
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        ExcelExportService._create_summary_sheet(wb, investigation, properties, companies, legal_queries)
        ExcelExportService._create_properties_sheet(wb, properties)
        ExcelExportService._create_companies_sheet(wb, companies)
        ExcelExportService._create_legal_queries_sheet(wb, legal_queries)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    @staticmethod
    def _create_summary_sheet(
        wb: Workbook,
        investigation: Investigation,
        properties: List[Property],
        companies: List[Company],
        legal_queries: List[LegalQuery],
    ) -> None:
        """Create summary sheet with investigation overview"""
        ws = wb.create_sheet("Resumo", 0)
        
        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Title
        ws["A1"] = "Relatório de Investigação - AgroADB"
        ws["A1"].font = Font(bold=True, size=16, color="4472C4")
        ws.merge_cells("A1:B1")
        
        # Investigation info
        row = 3
        data = [
            ("ID da Investigação:", investigation.id),
            ("Nome do Alvo:", investigation.target_name),
            ("CPF/CNPJ:", investigation.target_cpf_cnpj or "N/A"),
            ("Status:", investigation.status.value if hasattr(investigation.status, 'value') else str(investigation.status)),
            ("Prioridade:", investigation.priority),
            ("Data de Criação:", investigation.created_at.strftime("%d/%m/%Y %H:%M") if investigation.created_at else "N/A"),
            ("Última Atualização:", investigation.updated_at.strftime("%d/%m/%Y %H:%M") if investigation.updated_at else "N/A"),
            ("", ""),
            ("Resultados Encontrados:", ""),
            ("Propriedades:", len(properties)),
            ("Empresas:", len(companies)),
            ("Consultas Legais:", len(legal_queries)),
        ]
        
        for label, value in data:
            ws[f"A{row}"] = label
            ws[f"B{row}"] = value
            if label:
                ws[f"A{row}"].font = Font(bold=True)
            row += 1
        
        # Description section
        if investigation.target_description:
            row += 1
            ws[f"A{row}"] = "Descrição:"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            ws[f"A{row}"] = investigation.target_description
            ws.merge_cells(f"A{row}:B{row}")
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
        
        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

    @staticmethod
    def _create_properties_sheet(wb: Workbook, properties: List[Property]) -> None:
        """Create properties sheet"""
        ws = wb.create_sheet("Propriedades")
        
        # Header style
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = [
            "ID",
            "CAR",
            "Nome da Propriedade",
            "Área (ha)",
            "Município",
            "UF",
            "Fonte",
            "Data de Cadastro",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_num, prop in enumerate(properties, 2):
            ws.cell(row=row_num, column=1, value=prop.id)
            ws.cell(row=row_num, column=2, value=prop.car_number or "N/A")
            ws.cell(row=row_num, column=3, value=prop.property_name or "N/A")
            ws.cell(row=row_num, column=4, value=prop.area_hectares or 0)
            ws.cell(row=row_num, column=5, value=prop.city or "N/A")
            ws.cell(row=row_num, column=6, value=prop.state or "N/A")
            ws.cell(row=row_num, column=7, value=prop.data_source or "N/A")
            ws.cell(row=row_num, column=8, value=prop.created_at.strftime("%d/%m/%Y") if prop.created_at else "N/A")
        
        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

    @staticmethod
    def _create_companies_sheet(wb: Workbook, companies: List[Company]) -> None:
        """Create companies sheet"""
        ws = wb.create_sheet("Empresas")
        
        # Header style
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = [
            "ID",
            "CNPJ",
            "Razão Social",
            "Nome Fantasia",
            "Status",
            "Natureza Jurídica",
            "Município",
            "UF",
            "Data de Cadastro",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_num, company in enumerate(companies, 2):
            ws.cell(row=row_num, column=1, value=company.id)
            ws.cell(row=row_num, column=2, value=company.cnpj or "N/A")
            ws.cell(row=row_num, column=3, value=company.corporate_name or "N/A")
            ws.cell(row=row_num, column=4, value=company.trade_name or "N/A")
            ws.cell(row=row_num, column=5, value=company.status or "N/A")
            ws.cell(row=row_num, column=6, value=company.legal_nature or "N/A")
            ws.cell(row=row_num, column=7, value=company.city or "N/A")
            ws.cell(row=row_num, column=8, value=company.state or "N/A")
            ws.cell(row=row_num, column=9, value=company.created_at.strftime("%d/%m/%Y") if company.created_at else "N/A")
        
        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18

    @staticmethod
    def _create_legal_queries_sheet(wb: Workbook, legal_queries: List[LegalQuery]) -> None:
        """Create legal queries sheet"""
        ws = wb.create_sheet("Consultas Legais")
        
        # Header style
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = [
            "ID",
            "Provedor",
            "Tipo de Consulta",
            "Resultados",
            "Data da Consulta",
            "Parâmetros",
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data rows
        for row_num, query in enumerate(legal_queries, 2):
            ws.cell(row=row_num, column=1, value=query.id)
            ws.cell(row=row_num, column=2, value=query.provider or "N/A")
            ws.cell(row=row_num, column=3, value=query.query_type or "N/A")
            ws.cell(row=row_num, column=4, value=query.result_count or 0)
            ws.cell(row=row_num, column=5, value=query.created_at.strftime("%d/%m/%Y %H:%M") if query.created_at else "N/A")
            ws.cell(row=row_num, column=6, value=str(query.query_params) if query.query_params else "N/A")
        
        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Auto-adjust column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 40

    @staticmethod
    def generate_investigation_csv(
        investigation: Investigation,
        properties: List[Property],
        companies: List[Company],
        legal_queries: List[LegalQuery],
    ) -> BytesIO:
        """
        Generate CSV file with investigation data
        
        Args:
            investigation: Investigation object
            properties: List of properties found
            companies: List of companies found
            legal_queries: List of legal queries executed
            
        Returns:
            BytesIO: CSV file in memory
        """
        output = BytesIO()
        
        # Prepare data for pandas DataFrame
        data = {
            "ID Investigação": [investigation.id],
            "Nome do Alvo": [investigation.target_name],
            "CPF/CNPJ": [investigation.target_cpf_cnpj or "N/A"],
            "Status": [investigation.status.value if hasattr(investigation.status, 'value') else str(investigation.status)],
            "Prioridade": [investigation.priority],
            "Propriedades Encontradas": [len(properties)],
            "Empresas Encontradas": [len(companies)],
            "Consultas Legais": [len(legal_queries)],
            "Total de Resultados": [sum(q.result_count or 0 for q in legal_queries)],
            "Data de Criação": [investigation.created_at.strftime("%d/%m/%Y %H:%M") if investigation.created_at else "N/A"],
            "Última Atualização": [investigation.updated_at.strftime("%d/%m/%Y %H:%M") if investigation.updated_at else "N/A"],
        }
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Write to CSV
        df.to_csv(output, index=False, encoding="utf-8-sig")
        output.seek(0)
        
        return output
